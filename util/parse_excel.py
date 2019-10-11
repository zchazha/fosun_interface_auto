from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font

class ParseExcel:
    # 初始化表格
    def __init__(self,excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = load_workbook(excel_file_path)
        self.ws = self.wb.active
    # 通过sheet名获取sheet表
    def get_sheet_by_name(self,sheet_name):
        self.ws = self.wb.get_sheet_by_name(sheet_name)
        # sheet = self.wb[sheet_name]  # 方式二也可以
        return self.ws
    # 通过索引获取sheet表
    def get_sheet_by_index(self,index):
        sheet_name = self.wb.sheetnames[index-1]
        self.ws = self.wb[sheet_name]
        return self.ws
    # 获取所有的sheet名,列表格式
    def get_all_sheet_names(self):
        return self.wb.sheetnames
    # 获取单元格数据
    def get_cell_value(self,row_no,col_no,sheet_name=None):
        if sheet_name == None:
            return self.ws.cell(row=row_no, column=col_no).value
        return self.wb[sheet_name].cell(row=row_no, column=col_no).value
    # 获取单元格对象
    def get_cell_obj(self,row_no,col_no,sheet_name=None):
        if sheet_name == None:
            return self.ws.cell(row=row_no, column=col_no)
        return self.wb[sheet_name].cell(row=row_no, column=col_no)
    # 获取最大行号
    def get_max_row_num(self,sheet_name=None):
        if sheet_name == None:
            return self.ws.max_row
        return self.get_sheet_by_name(sheet_name).max_row
    # 获取最大列号
    def get_max_col_num(self,sheet_name=None):
        if sheet_name == None:
            return self.ws.max_column
        return self.get_sheet_by_name(sheet_name).max_column
    # 获取最小行号，默认为1
    def get_min_row_num(self,sheet_name=None):
        if sheet_name == None:
            return self.ws.min_row
        return self.get_sheet_by_name(sheet_name).min_row
    # 获取最小列号，默认为1
    def get_min_col_num(self,sheet_name=None):
        if sheet_name == None:
            return self.ws.min_column
        return self.get_sheet_by_name(sheet_name).min_column
    # 获取某行的值,rows中的行以0开始
    def get_some_row_value(self,row_no,sheet_name=None):
        row_value = []
        if sheet_name is not None:
            self.get_sheet_by_name(sheet_name)
        for i in list(self.ws.rows)[row_no-1]:
            row_value.append(i.value)
        return row_value
    # 获取某列的值,columns中的行以0开始
    def get_some_col_value(self,col_no,sheet_name=None):
        col_value = []
        if sheet_name is not None:
            self.get_sheet_by_name(sheet_name)
        for i in list(self.ws.columns)[col_no-1]:
            col_value.append(i.value)
        return col_value
    #保存单元格
    def save_excel(self):
        self.wb.save(self.excel_file_path)
    #单元格写入内容
    def write_cell_value(self,row_no,col_no,value,style=None,sheet_name=None):
        if sheet_name is not None:
            self.get_sheet_by_name(sheet_name)
        if style == None:
            style = colors.BLACK
        elif style.upper() == "RED":
            style = colors.RED
        elif style.upper() == "GREEN":
            style = colors.GREEN
        self.ws.cell(row=row_no, column=col_no).font=Font(color=style)
        self.ws.cell(row=row_no,column=col_no).value = value
        self.save_excel()
        return True

if __name__ == "__main__":
    pe = ParseExcel("sample_demo.xlsx")
    # ws = pe.get_sheet_by_name('表1')
    # print(pe.get_cell_value(1,1,'表1'))
    # print(pe.get_cell_obj(1,1,'表1'))
    # print(pe.get_cell_obj(1,1))
    pe.write_cell_value(10,10,"nihao",style="red",sheet_name="表1")
    pe.write_cell_value(10,11,"钉钉",style="green",sheet_name="表1")
